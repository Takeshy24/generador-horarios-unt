"""
Generación de Excel del horario con openpyxl.

GET /api/horario/excel/ciclo/{ciclo_num}?semestre_id=X    — grilla + tabla de carga del ciclo
GET /api/horario/excel/docente/{docente_id}?semestre_id=X — horario personal del docente
GET /api/horario/excel/completo?semestre_id=X             — libro completo con todos los ciclos
"""

from io import BytesIO
from datetime import date as date_type

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.db import get_db
from app.api.auth_routes import get_current_user
from app.models import Docente, Semestre
from app.api.horario import _get_bloques_db

router = APIRouter(prefix="/horario/excel", tags=["excel"])

DIAS = ["LUN", "MAR", "MIE", "JUE", "VIE", "SAB"]
DIA_LABELS = {
    "LUN": "Lunes", "MAR": "Martes", "MIE": "Miércoles",
    "JUE": "Jueves", "VIE": "Viernes", "SAB": "Sábado",
}
HORAS = [7, 8, 9, 10, 11, 12, 14, 15, 16, 17, 18, 19]

CICLO_ROMANO = {
    1: "I", 2: "II", 3: "III", 4: "IV", 5: "V",
    6: "VI", 7: "VII", 8: "VIII", 9: "IX", 10: "X",
}

_BG_COLORS = [
    "DBEAFE", "D1FAE5", "EDE9FE", "FEF3C7", "FCE7F3",
    "E0E7FF", "FFEDD5", "CCFBF1", "FEE2E2", "E0F2FE",
]
_TIPO_COLORS = {"T": "1D4ED8", "P": "065F46", "L": "5B21B6"}
_DEPT_FILLS  = [
    "FFFDD0", "D6EAF8", "D5F5E3", "FDEDEC",
    "F0E6FA", "FEF9E7", "D0ECE7", "F2F3F4",
]

_DIA_COL  = {dia: i + 2 for i, dia in enumerate(DIAS)}  # LUN→col2 … SAB→col7
_TITLE_ROW  = 1
_HEADER_ROW = 2
_LUNCH_ROW  = 9   # between hora 12 (row 8) and hora 14 (row 10)


def _hora_to_row(hora: int) -> int:
    if 7 <= hora <= 12:
        return hora - 7 + _HEADER_ROW + 1   # 7→3 … 12→8
    if 14 <= hora <= 19:
        return hora - 14 + _LUNCH_ROW + 1   # 14→10 … 19→15
    raise ValueError(f"Hora {hora} fuera de rango")


def _border(color: str = "CCCCCC", width: str = "thin") -> Border:
    s = Side(style=width, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ── Hoja de grilla ────────────────────────────────────────────────────────────

def _build_horario_sheet(ws, bloques: list[dict], title: str, semestre_codigo: str) -> None:
    ws.sheet_view.showGridLines = True

    # Column widths
    ws.column_dimensions["A"].width = 7
    for i in range(len(DIAS)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 24

    # Title row
    ws.merge_cells(start_row=_TITLE_ROW, start_column=1, end_row=_TITLE_ROW, end_column=7)
    tc = ws.cell(row=_TITLE_ROW, column=1,
                 value=f"{title}  ·  Semestre {semestre_codigo}  ·  {date_type.today().strftime('%d/%m/%Y')}")
    tc.font = Font(bold=True, size=13, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="17499E")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[_TITLE_ROW].height = 26

    # Day headers
    ws.row_dimensions[_HEADER_ROW].height = 22
    hora_hdr = ws.cell(row=_HEADER_ROW, column=1, value="Hora")
    hora_hdr.font = Font(bold=True, size=9, color="17499E")
    hora_hdr.fill = PatternFill("solid", fgColor="EEF2FF")
    hora_hdr.alignment = Alignment(horizontal="center", vertical="center")
    hora_hdr.border = _border()

    for i, dia in enumerate(DIAS):
        hdr = ws.cell(row=_HEADER_ROW, column=i + 2, value=DIA_LABELS[dia])
        hdr.font = Font(bold=True, size=10, color="17499E")
        hdr.fill = PatternFill("solid", fgColor="EEF2FF")
        hdr.alignment = Alignment(horizontal="center", vertical="center")
        hdr.border = _border()

    # Hour label cells + blank day cells
    for hora in HORAS:
        row = _hora_to_row(hora)
        ws.row_dimensions[row].height = 44

        lbl = ws.cell(row=row, column=1, value=f"{hora:02d}:00")
        lbl.font = Font(size=8, color="666666")
        lbl.fill = PatternFill("solid", fgColor="F5F7FA")
        lbl.alignment = Alignment(horizontal="center", vertical="center")
        lbl.border = _border()

        for col in range(2, 8):
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill("solid", fgColor="FFFFFF")
            c.border = _border()

    # Lunch separator
    ws.row_dimensions[_LUNCH_ROW].height = 13
    ws.merge_cells(start_row=_LUNCH_ROW, start_column=1, end_row=_LUNCH_ROW, end_column=7)
    lunch = ws.cell(row=_LUNCH_ROW, column=1, value="— Almuerzo —")
    lunch.font = Font(italic=True, size=7.5, color="AAAAAA")
    lunch.fill = PatternFill("solid", fgColor="FFF7ED")
    lunch.alignment = Alignment(horizontal="center", vertical="center")

    # Place schedule blocks
    occupied: set[tuple[int, int]] = set()
    for bloque in bloques:
        dia = bloque["dia"]
        if dia not in _DIA_COL:
            continue
        start_h = int(bloque["hora_inicio"][:2])
        end_h   = int(bloque["hora_fin"][:2])
        try:
            start_row = _hora_to_row(start_h)
            end_row   = _hora_to_row(end_h)
        except ValueError:
            continue

        col = _DIA_COL[dia]
        if (start_row, col) in occupied:
            continue
        for r in range(start_row, end_row + 1):
            occupied.add((r, col))

        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=col,
                           end_row=end_row, end_column=col)

        cell = ws.cell(row=start_row, column=col)

        comp   = bloque["componente"]
        curso  = comp["seccion"]["curso"]
        tipo   = comp["tipo"]
        doc    = comp["docente"]
        sec    = comp["seccion"]["letra"]
        aula   = bloque["aula"]["codigo"]

        tipo_label = {"T": "Teoría", "P": "Práctica", "L": "Lab"}.get(tipo, tipo)
        doc_name   = doc["nombre"].split(",")[0].strip() if doc else "Sin docente"

        cell.value = f"[{tipo_label}]\n{curso['nombre']}\n{doc_name}\nSec.{sec} · {aula}"
        cell.fill  = PatternFill("solid", fgColor=_BG_COLORS[curso["id"] % len(_BG_COLORS)])
        cell.font  = Font(size=8, color=_TIPO_COLORS.get(tipo, "333333"))
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(
            left   = Side(style="medium", color=_TIPO_COLORS.get(tipo, "666666")),
            right  = Side(style="thin",   color="CCCCCC"),
            top    = Side(style="thin",   color="CCCCCC"),
            bottom = Side(style="thin",   color="CCCCCC"),
        )

    # Footer
    footer_row = _hora_to_row(19) + 1
    ws.row_dimensions[footer_row].height = 13
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=7)
    footer = ws.cell(row=footer_row, column=1,
                     value="Universidad Nacional de Trujillo  ·  Escuela de Ingeniería de Sistemas")
    footer.font = Font(size=7, italic=True, color="AAAAAA")
    footer.alignment = Alignment(horizontal="center")


# ── Hoja de carga docente ─────────────────────────────────────────────────────

def _build_carga_sheet(ws, bloques: list[dict], ciclo_num: int, semestre: Semestre) -> None:
    parts = semestre.codigo.split("-")
    año, sem_num = (parts[0], parts[1]) if len(parts) == 2 else (semestre.codigo, "")
    ciclo_romano = CICLO_ROMANO.get(ciclo_num, str(ciclo_num))

    # Institutional header
    for row, text, size in [
        (1, "UNIVERSIDAD NACIONAL DE TRUJILLO — FACULTAD DE INGENIERÍA", 11),
        (2, "ESCUELA DE INGENIERÍA DE SISTEMAS", 10),
        (3, "DISTRIBUCIÓN DE CARGA DOCENTE", 10),
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=size, color="17499E" if row == 3 else "000000")
        c.fill = PatternFill("solid", fgColor="EEF2FF") if row == 3 else PatternFill("solid", fgColor="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 18

    # Metadata
    meta = [
        (5, 1, "ESCUELA:", 2, "Ingeniería de Sistemas"),
        (5, 5, "CICLO:",   6, ciclo_romano),
        (6, 1, "AÑO:",     2, año),
        (6, 5, "SEMESTRE:", 6, sem_num),
        (7, 1, "INICIO:",  2, semestre.fecha_inicio.strftime("%d/%m/%Y")),
        (7, 5, "TÉRMINO:", 6, semestre.fecha_fin.strftime("%d/%m/%Y")),
    ]
    for row, lbl_col, lbl, val_col, val in meta:
        ws.cell(row=row, column=lbl_col, value=lbl).font = Font(bold=True, size=9)
        ws.cell(row=row, column=val_col, value=val).font = Font(size=9)
        ws.row_dimensions[row].height = 15

    # Column headers
    COLS       = ["Nº", "PROFESOR", "ASIGNATURA", "T", "P", "L", "G", "T. HORAS", "DEPARTAMENTO"]
    COL_WIDTHS = [5,    30,          32,           5,   5,   5,   5,   10,          24           ]
    HDR_ROW = 9
    ws.row_dimensions[HDR_ROW].height = 26

    for i, (label, width) in enumerate(zip(COLS, COL_WIDTHS)):
        col = i + 1
        ws.column_dimensions[get_column_letter(col)].width = width
        is_text = col in (2, 3)
        hdr = ws.cell(row=HDR_ROW, column=col, value=label)
        hdr.font = Font(bold=True, size=9,
                        color="1A1A1A" if is_text else "FFFFFF")
        hdr.fill = PatternFill("solid", fgColor="F5F0B0" if is_text else "1E3A6E")
        hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hdr.border = _border("999999")

    # Data rows
    seen: set[int] = set()
    carga: dict = {}
    for b in bloques:
        comp = b["componente"]
        cid = comp["id"]
        if cid in seen:
            continue
        seen.add(cid)
        doc = comp["docente"]
        if not doc:
            continue
        curso = comp["seccion"]["curso"]
        key = (doc["nombre"], doc.get("departamento", "—"), curso["nombre"])
        if key not in carga:
            carga[key] = {"T": 0, "P": 0, "L": 0, "G": 0}
        carga[key][comp["tipo"]] = comp["horas_semanales"]

    rows = sorted(
        [{"docente": k[0], "dpto": k[1], "curso": k[2],
          "T": v["T"], "P": v["P"], "L": v["L"], "G": v["G"],
          "total": v["T"] + v["P"] + v["L"] + v["G"]}
         for k, v in carga.items()],
        key=lambda r: r["docente"],
    )

    depts = sorted({r["dpto"] for r in rows})
    dept_fill = {d: _DEPT_FILLS[i % len(_DEPT_FILLS)] for i, d in enumerate(depts)}
    brdr = _border("CCCCCC")

    for i, row in enumerate(rows):
        excel_row = HDR_ROW + 1 + i
        ws.row_dimensions[excel_row].height = 17
        fill = PatternFill("solid", fgColor=dept_fill.get(row["dpto"], "FFFFFF"))
        values = [
            i + 1, row["docente"], row["curso"],
            row["T"] or "—", row["P"] or "—", row["L"] or "—", row["G"] or "—",
            row["total"], row["dpto"],
        ]
        for j, val in enumerate(values):
            col = j + 1
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.fill = fill
            cell.border = brdr
            cell.font = Font(size=8.5)
            cell.alignment = Alignment(
                horizontal="left" if col in (2, 3, 9) else "center",
                vertical="center",
                wrap_text=True,
            )

    # Footer
    footer_row = HDR_ROW + 1 + len(rows) + 1
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=9)
    footer = ws.cell(row=footer_row, column=1,
                     value=f"Sistema de Generación de Horarios · Escuela de Ingeniería de Sistemas · {date_type.today().strftime('%d/%m/%Y')}")
    footer.font = Font(size=7, italic=True, color="AAAAAA")
    footer.alignment = Alignment(horizontal="center")


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/ciclo/{ciclo_num}")
async def excel_ciclo(
    ciclo_num: int,
    semestre_id: int = Query(...),
    user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Excel con la grilla del ciclo (hoja 1) + tabla de carga docente (hoja 2)."""
    sem_res = await db.execute(select(Semestre).where(Semestre.id == semestre_id))
    semestre = sem_res.scalar_one_or_none()
    if not semestre:
        raise HTTPException(status_code=404, detail="Semestre no encontrado")

    bloques = await _get_bloques_db(db, semestre_id, ciclo=ciclo_num)
    if not bloques:
        raise HTTPException(status_code=404, detail=f"No hay bloques para el ciclo {ciclo_num}")

    num_romano = CICLO_ROMANO.get(ciclo_num, str(ciclo_num))

    wb = Workbook()
    ws_h = wb.active
    ws_h.title = f"Horario Ciclo {num_romano}"
    _build_horario_sheet(ws_h, bloques, f"Horario de Clases — Ciclo {num_romano}", semestre.codigo)

    ws_c = wb.create_sheet(title="Carga Docente")
    _build_carga_sheet(ws_c, bloques, ciclo_num, semestre)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"horario_ciclo{ciclo_num}_{semestre.codigo}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/docente/{docente_id}")
async def excel_docente(
    docente_id: int,
    semestre_id: int = Query(...),
    user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Excel con el horario personal del docente."""
    sem_res = await db.execute(select(Semestre).where(Semestre.id == semestre_id))
    semestre = sem_res.scalar_one_or_none()
    if not semestre:
        raise HTTPException(status_code=404, detail="Semestre no encontrado")

    doc_res = await db.execute(select(Docente).where(Docente.id == docente_id))
    docente = doc_res.scalar_one_or_none()
    if not docente:
        raise HTTPException(status_code=404, detail="Docente no encontrado")

    bloques = await _get_bloques_db(db, semestre_id, docente_id=docente_id)
    if not bloques:
        raise HTTPException(status_code=404, detail="No hay bloques para este docente")

    wb = Workbook()
    ws = wb.active
    ws.title = "Mi Horario"
    _build_horario_sheet(
        ws, bloques,
        f"Horario Personal — {docente.nombre_completo}",
        semestre.codigo,
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe = docente.nombre_completo.replace(",", "").replace(" ", "_")[:30]
    filename = f"horario_{safe}_{semestre.codigo}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/completo")
async def excel_completo(
    semestre_id: int = Query(...),
    user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Excel completo: un par de hojas (Horario + Carga) por cada ciclo con bloques."""
    sem_res = await db.execute(select(Semestre).where(Semestre.id == semestre_id))
    semestre = sem_res.scalar_one_or_none()
    if not semestre:
        raise HTTPException(status_code=404, detail="Semestre no encontrado")

    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet
    sheets_added = 0

    for ciclo_num in range(1, 11):
        bloques = await _get_bloques_db(db, semestre_id, ciclo=ciclo_num)
        if not bloques:
            continue
        num_romano = CICLO_ROMANO.get(ciclo_num, str(ciclo_num))

        ws_h = wb.create_sheet(title=f"Horario C{num_romano}")
        _build_horario_sheet(
            ws_h, bloques,
            f"Horario de Clases — Ciclo {num_romano}",
            semestre.codigo,
        )
        ws_c = wb.create_sheet(title=f"Carga C{num_romano}")
        _build_carga_sheet(ws_c, bloques, ciclo_num, semestre)
        sheets_added += 1

    if sheets_added == 0:
        raise HTTPException(status_code=404, detail="No hay horario generado para este semestre")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"horario_completo_{semestre.codigo}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
